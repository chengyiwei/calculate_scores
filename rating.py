import os
import openpyxl
import re
import math
import requests
from bs4 import BeautifulSoup

# 获取学生信息列表
def get_stu_info():
    # 获取当前文件所在的文件夹路径
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\information.xlsx"
    workbook  = openpyxl.load_workbook(path) 
    sheet = workbook.active # 打开活动表单
    stu_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1 # 获取学生数量
    stu = {} # 学生信息字典
    stu_list = {} # 学生信息字典集合
    for i in range(2, int(stu_num) + 2):
        stu.clear()
        name = sheet.cell(row=i, column=2).value
        if (name == None):
            continue
        stu['luogu_name'] = sheet.cell(row=i, column=3).value
        stu['at_id'] = sheet.cell(row=i, column=4).value
        stu['cf_id'] = sheet.cell(row=i, column=5).value
        stu_list[name] = stu.copy()
    return stu_list


# 获取学生总分数列表
def get_stu_all_rating():
    stu_info = get_stu_info()
    # 获取当前文件所在的文件夹路径
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\rating.xlsx"
    workbook  = openpyxl.load_workbook(path) 
    sheet = workbook.active # 打开活动表单
    student_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1 # 获取学生数量
    cur_stu = {} # 学生信息字典
    stu_dic = {} # 学生信息字典序列
    for i in range(2, int(student_num) + 2):
        cur_stu.clear()
        name = sheet.cell(row=i, column=1).value
        if name not in stu_info.keys(): # 如果学生信息表中没有该学生信息，则跳过
            continue
        cur_stu['rating'] = sheet.cell(row=i, column=2).value
        cur_stu['at_rating'] = sheet.cell(row=i, column=3).value
        cur_stu['cf_rating'] = sheet.cell(row=i, column=4).value
        stu_dic[name] = cur_stu.copy()
    workbook.save(path)
    return stu_dic


# 获取比赛列表
# 返回列表 [(year, id)]
def get_contest_list():
    # 获取当前文件所在的文件夹路径
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    directory = current_folder_path + "\\files\\contests"
    contest_list = os.listdir(directory)
    ret_list = []
    for file_name in contest_list:
        year = int(file_name.split('.')[0].split('#')[0])
        id = int(file_name.split('.')[0].split('#')[1])
        ret_list.append((year, id))
    # 按照年份和 id 从大到小排序
    ret_list.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return ret_list
        

# 获取比赛信息
# 返回字典 {name: rank} 无序 如果找不到比赛则返回空字典
def get_contest_info(x):
    stu_info = get_stu_info()
    # stu_real_name : luogu_name -> name 建立反向字典
    stu_luogu_name = {}
    for stu in stu_info:
        stu_luogu_name[stu_info[stu]['luogu_name']] = stu

    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\contests\\" + str(x) + ".xlsx"

    try:
        workbook  = openpyxl.load_workbook(path)
    except:
        print("找不到名称为 " + str(x) + " 的比赛")
        return {}
    sheet = workbook.active
    stu_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1

    stu_list = []
    for i in range(2, int(stu_num) + 2):
        luogu_name = sheet.cell(row=i, column=2).value
        if luogu_name not in stu_luogu_name: # 参赛人员不在我们要统计的名单中
            continue
        name = stu_luogu_name[luogu_name]
        rank = int(sheet.cell(row=i, column=1).value)
        stu_list.append({'name': name, 'rank': rank})
    
    sorted(stu_list, key=lambda x: x['rank']) # 按照 rank 排序
    i, j = 0, 0
    while i < len(stu_list):
        j = i
        while j < len(stu_list) and stu_list[j]['rank'] == stu_list[i]['rank']:
            j += 1
        for k in range(i, j):
            stu_list[k]['rank'] = i + 1
        i = j
    
    stu_dic = {}
    for stu in stu_list:
        stu_dic[stu['name']] = stu['rank']
    
    return stu_dic



# 获取某个学生的比赛记录
# 输入学生姓名 name
# 返回列表 [{contest_name: , rank: , new_rating: }]
def get_member_record(name):
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\member_record\\" + name + ".xlsx"
    try:
        workbook  = openpyxl.load_workbook(path)
    except:
        print("找不到名称为 " + str(name) + " 的比赛记录")
        return []
    sheet = workbook.active
    contest_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1
    if contest_num == 0:
        print(str(name) + " 没有参加比赛")
        return []
    contest_list = []
    for i in range(2, int(contest_num) + 2):
        contest = {}
        contest['contest_name'] = sheet.cell(row=i, column=1).value
        contest['rank'] = sheet.cell(row=i, column=2).value
        contest['new_rating'] = sheet.cell(row=i, column=3).value
        contest_list.append(contest)
    return contest_list


# 计算第 x 场比赛的分数变化
# 读入比赛名称 contest_name
# 返回列表 [{name: ,rank:, old_rating:, delta:, new_rating}]
def calc_rating_change(contest_name, is_rated=False):
    RATEDBOUND = 4000

    contest_info = get_contest_info(contest_name)
    if contest_info == {}:
        return [] # 没有找到比赛
    stu_all_rating = get_stu_all_rating()
    stu_num = len(contest_info)

    stu_list = []
    for name in contest_info: # 初始化列表
        cur_stu = {}
        cur_stu['name'] = name
        cur_stu['rank'] = contest_info[name]
        cur_stu['old_rating'] = stu_all_rating[name]['rating']
        stu_list.append(cur_stu.copy())
    sorted(stu_list, key=lambda x: x['rank']) # 按照 rank 排序
    
    if is_rated == False: # 非 rated 比赛
        for i in range(stu_num):
            stu_list[i]['delta'] = 0
            stu_list[i]['new_rating'] = stu_list[i]['old_rating']
        return stu_list

    perf = [[0 for i in range(stu_num)] for j in range(stu_num)]
    for i in range(stu_num):
        for j in range(stu_num):
            perf[i][j] = 1 / (1 + math.pow(10, (stu_list[j]['old_rating'] - stu_list[i]['old_rating']) / 400))

    for i in range(stu_num):
        seed = 1
        for j in range(stu_num):
            if j == i:
                continue
            seed += perf[i][j]
        stu_list[i]['seed'] = seed

    i, j = 0, 0
    while i < stu_num:
        j = i
        while j < stu_num and stu_list[j]['rank'] == stu_list[i]['rank']:
            j += 1
        avg_rank = (i + j + 1) / 2
        for k in range(i, j):
            stu_list[k]['m'] = math.sqrt(stu_list[k]['seed'] * avg_rank)
        i = j

    def calc_rating(cur_stu):
        Le, Ri = 1, RATEDBOUND
        for x in range(20):
            mid = (Le + Ri) / 2
            new_seed = 1
            for j in range(stu_num):
                Pji = 1 / (1 + math.pow(10, (mid - stu_list[j]['old_rating']) / 400))
                new_seed += Pji
            if new_seed > cur_stu['m']:
                Le = mid
            else:
                Ri = mid
        return int((Le + Ri) / 2)
    
    for i in range(stu_num):
        Ri = calc_rating(stu_list[i])
        now_rating = int((2 * stu_list[i]['old_rating'] + Ri) / 3)
        stu_list[i]['delta'] = now_rating - stu_list[i]['old_rating']
    
    # 第一次微调
    sum_delta = 0
    for i in range(stu_num):
        sum_delta += stu_list[i]['delta']
    inc = (-1 - sum_delta) // stu_num
    for i in range(stu_num):
        stu_list[i]['delta'] += inc
    
    # 第二次微调
    sum_delta = 0
    for i in range(stu_num):
        sum_delta += stu_list[i]['delta']
    inc = 0
    while sum_delta < 0:
        inc += 1
        sum_delta += stu_num
    for i in range(stu_num):
        stu_list[i]['delta'] += inc
    

    # 更新 new_rating，删除 seed 和 m
    for i in range(stu_num):
        stu_list[i]['new_rating'] = stu_list[i]['old_rating'] + stu_list[i]['delta']
        stu_list[i].pop('seed')
        stu_list[i].pop('m')
    
    return stu_list


# 获取学生比赛分数变化 
# 输入比赛名称 contest_name
# 返回字典 {name: {rank: , old_rating: , delta: , new_rating: }}
def get_rating_change_info(contest_name):
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\rating_change\\" + str(contest_name) + ".xlsx"
    try:
        workbook  = openpyxl.load_workbook(path)
    except:
        print("找不到名称为 " + str(contest_name) + " 的比赛")
        return {}
    sheet = workbook.active
    stu_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1
    stu_dic = {}
    for i in range(2, int(stu_num) + 2):
        cur_stu = {}
        cur_stu['rank'] = sheet.cell(row=i, column=1).value
        cur_stu['name'] = sheet.cell(row=i, column=2).value
        cur_stu['old_rating'] = sheet.cell(row=i, column=3).value
        cur_stu['delta'] = sheet.cell(row=i, column=4).value
        cur_stu['new_rating'] = sheet.cell(row=i, column=5).value
        name = cur_stu['name']
        stu_dic[name] = cur_stu.copy()
    return stu_dic


# 导入 contest_name 比赛，并产生日志
def update_contest(contest_name):
    print("正在更新名为 " + str(contest_name) + " 的比赛 ...")
    
    is_rated = input("是否为 rated 比赛？(y/n) ")
    if is_rated == "y" or is_rated == "Y" or is_rated == "yes" or is_rated == "Yes" or is_rated == "YES":
        is_rated = True
    else:
        is_rated = False
    stu_list = calc_rating_change(contest_name, is_rated)

    if stu_list == []:
        print("找不到名为 " + str(contest_name) + " 的比赛")
        return
    
    def save_rating_change(stu_list):
        # 获取当前文件所在的文件夹路径
        current_file_path = os.path.abspath(__file__)
        current_folder_path = os.path.dirname(current_file_path)
        os.chdir(current_folder_path)
        
        # 创建一个新的工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "rating_change"

        stu_num = len(stu_list)
        sheet.cell(row=1, column=1).value = "rank"
        sheet.cell(row=1, column=2).value = "name"
        sheet.cell(row=1, column=3).value = "old_rating"
        sheet.cell(row=1, column=4).value = "delta"
        sheet.cell(row=1, column=5).value = "new_rating"

        for i in range(2, stu_num + 2):
            sheet.cell(row=i, column=1).value = stu_list[i - 2]['rank']
            sheet.cell(row=i, column=2).value = stu_list[i - 2]['name']
            sheet.cell(row=i, column=3).value = stu_list[i - 2]['old_rating']
            sheet.cell(row=i, column=4).value = stu_list[i - 2]['delta']
            sheet.cell(row=i, column=5).value = stu_list[i - 2]['new_rating']
        
        # 保存文件
        save_path = os.path.join(current_folder_path, "files" ,"rating_change", f"{contest_name}.xlsx")
        workbook.save(save_path)

    save_rating_change(stu_list)
    print("日志已保存至 rating_change 文件夹的 " + str(contest_name) + ".xlsx")

    stu_all_rating = get_stu_all_rating()
    for stu in stu_list:
        stu_all_rating[stu['name']]['rating'] = stu['new_rating']
    
    save_scores_xlsx(stu_all_rating)
    print("rating 已更新至 rating.xlsx")


# 创建所有学生的比赛记录
# 无返回值，直接保存
def create_all_stu_record():
    stu_info = get_stu_info()
    contest_list = get_contest_list()
    
    contest_record_list = []
    for contest in contest_list:
        contest_name = str(contest[0]) + "#" + str(contest[1])
        contest_record = get_rating_change_info(contest_name)
        contest_record_list.append({'name': contest_name, 'record': contest_record})

    def save_stu_record(name):
        stu_record = []
        for contest in contest_record_list:
            if name in contest['record']:
                stu_record.append({'name': contest['name'], 'rank': contest['record'][name]['rank'] , 'new_rating': contest['record'][name]['new_rating']})
            else:
                pass
        # 获取当前文件所在的文件夹路径
        current_file_path = os.path.abspath(__file__)
        current_folder_path = os.path.dirname(current_file_path)
        path = current_folder_path + "\\files\\member_record\\" + name + ".xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = name
        sheet.cell(row=1, column=1).value = "比赛名称"
        sheet.cell(row=1, column=2).value = "排名"
        sheet.cell(row=1, column=3).value = "新 rating"
        for i in range(2, len(stu_record) + 2):
            sheet.cell(row=i, column=1).value = stu_record[i - 2]['name']
            sheet.cell(row=i, column=2).value = stu_record[i - 2]['rank']
            sheet.cell(row=i, column=3).value = stu_record[i - 2]['new_rating']
        workbook.save(path)
        
    
    for name in stu_info:
        print("正在创建 " + name + " 的比赛记录 ... ")
        save_stu_record(name)
    


# 保存学生分数列表, 无返回值
def save_scores_xlsx(stu_all_rating):
    stu_list = []
    for x in stu_all_rating:
        stu = {}
        stu['name'] = x
        stu['rating'] = stu_all_rating[x]['rating']
        stu['at_rating'] = stu_all_rating[x]['at_rating']
        stu['cf_rating'] = stu_all_rating[x]['cf_rating']
        stu_list.append(stu.copy())

    stu_list.sort(key=lambda x: x['rating'], reverse=True) # 按照 rating 降序排序
    
    # 获取当前文件所在的文件夹路径
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\rating.xlsx"
    workbook  = openpyxl.load_workbook(path)
    sheet = workbook.active
    stu_num = stu_list.__len__()
    for i in range(2, stu_num + 2):
        sheet.cell(row=i, column=1).value = stu_list[i - 2]['name']
        sheet.cell(row=i, column=2).value = stu_list[i - 2]['rating']
        sheet.cell(row=i, column=3).value = stu_list[i - 2]['at_rating']
        sheet.cell(row=i, column=4).value = stu_list[i - 2]['cf_rating']

    workbook.save(path)


# 爬虫获取 at 的 rating
def get_at_rating():
    print("正在爬取 atcoder rating ...")
    stu_info = get_stu_info()
    stu_at_name = {} # at_id -> name
    stu_at_rating = {}

    for name in stu_info:
        at_id = stu_info[name]['at_id']
        if at_id == None:
            stu_at_rating[name] = -1
            continue
        stu_at_name[at_id] = name
    
    root_url = "https://atcoder.jp"
    
    def get_at_user_rating(at_id):
        url = root_url + "/users/" + at_id
        response = requests.get(url)
        if response.status_code != 200:
            return
    
        soup = BeautifulSoup(response.text, 'html.parser')
        th_tag = soup.find('th', string='Rating')
        if th_tag:
            rating_str = th_tag.find_next('td').text.strip()
        else:
            return 0 # 未参加比赛
        
        #使用正则表达式提取 rating_str 中的数字
        rating = int(re.findall(r'\d+', rating_str)[0])
        return rating
    
    for at_id in stu_at_name:
        name = stu_at_name[at_id]
        print("正在爬取 " + name + " 的 rating ... ")
        rating = get_at_user_rating(at_id)
        if rating == None:
            rating = -1
        stu_at_rating[name] = rating
    
    return stu_at_rating


# 更新 atcoder rating
def update_at_rating():
    print("正在更新 atcoder rating ...")
    stu_at_rating = get_at_rating()
    stu_all_rating = get_stu_all_rating()
    for name in stu_at_rating:
        stu_all_rating[name]['at_rating'] = stu_at_rating[name]
    save_scores_xlsx(stu_all_rating)
    print("atcoder rating 已更新至 rating.xlsx")


# 爬虫获取 cf 的 rating
def get_cf_rating():
    print("正在爬取 codeforces rating ...")
    stu_info = get_stu_info()
    stu_cf_name = {} # cf_id -> name
    stu_cf_rating = {}
    for name in stu_info:
        cf_id = stu_info[name]['cf_id']
        if cf_id == None:
            stu_cf_rating[name] = -1
            continue
        stu_cf_name[cf_id] = name
    
    root_url = "https://codeforces.com"
    def get_cf_user_rating(cf_id):
        if cf_id == None:
            return
        url = root_url + "/profile/" + cf_id
        response = requests.get(url)
        if response.status_code != 200:
            return
        soup = BeautifulSoup(response.text, 'html.parser')
        try:
            tag = soup.find(string=lambda text: text and "Contest rating:" in text)
            if tag:
                rating = tag.find_next('span').text.strip()
            else:
                rating = "0"
        except AttributeError:
            rating = "0"
        return int(rating)
    
    for cf_id in stu_cf_name:
        name = stu_cf_name[cf_id]
        print("正在爬取 " + name + " 的 rating ... ")
        rating = get_cf_user_rating(cf_id)
        if rating == None:
            rating = -1
        stu_cf_rating[name] = rating
    
    return stu_cf_rating


# 更新 codeforces rating
def update_cf_rating():
    print("正在更新 codeforces rating ...")
    stu_cf_rating = get_cf_rating()
    stu_all_rating = get_stu_all_rating()
    for name in stu_cf_rating:
        stu_all_rating[name]['cf_rating'] = stu_cf_rating[name]
    save_scores_xlsx(stu_all_rating)
    print("codeforces rating 已更新至 rating.xlsx")


# 强制改变 rating
def force_change_rating():
    print("正在强制改变 rating ...")
    stu_all_rating = get_stu_all_rating()

    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\force.xlsx"
    workbook  = openpyxl.load_workbook(path)
    sheet = workbook.active

    stu_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1
    stu_dic = {}
    for i in range(2, int(stu_num) + 2):
        name = sheet.cell(row=i, column=1).value
        rating = sheet.cell(row=i, column=2).value
        stu_dic[name] = rating
    for name in stu_dic:
        stu_all_rating[name]['rating'] = stu_dic[name]
    save_scores_xlsx(stu_all_rating)
    print("focre 的数据已更新至 rating.xlsx")


# 指令控制
def command(s):
    if s == "reset" or s == "re":
        force_change_rating()
        return 1
    lst = s.split(' ')
    if lst[0] == "upd" or lst[0] == "update":
        if lst[1] == "at" or lst[1] == "atcoder":
            update_at_rating()
            return 1
        if lst[1] == "cf" or lst[1] == "codeforces":
            update_cf_rating()
            return 1
        if lst[1] == "contest" or lst[1] == "c":
            update_contest(lst[2])
            return 1
    if lst[0] == "get":
        if lst[1] == "info" or lst[1] == "i":
            stu_info = get_stu_info()
            for x in stu_info:
                print(f"{x} : {stu_info[x]}")
            return 1
        if lst[1] == "rating" or lst[1] == "r":
            stu_all_rating = get_stu_all_rating()
            for x in stu_all_rating:
                print(f"{x} : {stu_all_rating[x]}")
            return 1
        if lst[1] == "at" or lst[1] == "atcoder":
            stu_at_rating = get_at_rating()
            for x in stu_at_rating:
                print(f"{x} : {stu_at_rating[x]}")
            return 1
        if lst[1] == "cf" or lst[1] == "codeforces":
            stu_cf_rating = get_cf_rating()
            for x in stu_cf_rating:
                print(f"{x} : {stu_cf_rating[x]}")
            return 1
        if lst[1] == "contest":
            stu_contest = get_rating_change_info(lst[2])
            for x in stu_contest:
                print(f"{x} : {stu_contest[x]}")
            return 1
    if lst[0] == "pre":
        x = lst[1]
        stu_list = calc_rating_change(x)
        for stu in stu_list:
            print(stu)
        return 1
    print("找不到指令")
    return 0



def contrl():
    s = input("请输入指令：")
    while command(s) == 0:
        s = input("请输入指令：")


def QQ_ask(s):
    HELP = "/info 输出atcoder和codeforces的id\n/at 输出atcoder的rating\n/cf 输出codeforces的rating\n/rating 输出训练赛的rating\n/contest [contest_name] 输出rating变化值\n/stu [name] 输出某个人的比赛记录\n"
    s = s[1:].strip()

    # 单命令查询

    if s == "help":
        return HELP

    if s == "info":
        stu_info = get_stu_info()
        res = ""
        for x in stu_info:
            res += f"{x} at_id: {stu_info[x]['at_id']} cf_id: {stu_info[x]['cf_id']}\n"
        return res

    if s == "at" or s == "AT":
        res = ""
        stu_all_rating = get_stu_all_rating()
        stu_list = []
        for x in stu_all_rating:
            stu_list.append({'name': x, 'at_rating': stu_all_rating[x]['at_rating']})
        stu_list.sort(key=lambda x: x['at_rating'], reverse=True)
        for x in stu_list:
            res += f"{x['name']} : {x['at_rating']}\n"
        return res
    
    if s == "cf" or s == "CF":
        res = ""
        stu_all_rating = get_stu_all_rating()
        stu_list = []
        for x in stu_all_rating:
            stu_list.append({'name': x, 'cf_rating': stu_all_rating[x]['cf_rating']})
        stu_list.sort(key=lambda x: x['cf_rating'], reverse=True)
        for x in stu_list:
            res += f"{x['name']} : {x['cf_rating']}\n"
        return res
    
    if s == "rating" or s == "Rating":
        res = ""
        stu_all_rating = get_stu_all_rating()
        stu_list = []
        for x in stu_all_rating:
            stu_list.append({'name': x, 'rating': stu_all_rating[x]['rating']})
        stu_list.sort(key=lambda x: x['rating'], reverse=True)
        for x in stu_list:
            res += f"{x['name']} : {x['rating']}\n"
        return res
    
    # 多命令查询
    list_s= s.split(" ")
    if list_s[0] == "contest":
        stu_rating_change = get_rating_change_info(list_s[1])
        if stu_rating_change == {}:
            return "没有找到该比赛"
        res = ""
        stu_list = []
        for x in stu_rating_change:
            stu_list.append({'name': x, 'rank': stu_rating_change[x]['rank'], 'delta': stu_rating_change[x]['delta']})
        stu_list.sort(key=lambda x: x['rank'])
        for x in stu_list:
            res += f"{x['rank']} : {x['name']}   {x['delta']}\n"
        return res

    if list_s[0] == "stu":
        stu_record = get_member_record(list_s[1])
        if stu_record == []:
            return "没有找到该学生的比赛记录"
        res = ""
        for x in stu_record:
            res += f"{x['contest_name']} : rk {x['rank']} -> {x['new_rating']}\n"
        return res

    return "error"

create_all_stu_record()